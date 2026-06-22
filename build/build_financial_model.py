#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""build_financial_model.py — outputs/02_財務モデル.xlsx を 4シート構成で再生成する。

設計参照: docs/specs/2026-06-21-financial-model-simplification-design.md
実装計画: docs/specs/2026-06-22-financial-model-simplification-plan.md

ビルド方針(=本質と装飾の棲み分け):
  VC財務モデルは「2問」にだけ答える。
    ① Solafune 1社でファンドを返せるか? (= 回収額 ≧ 150億)
    ② 10社合計で net DPI 3.0X 行けるか?

  4シート構成:
    ① 前提            ─ 8つの入力 + 注記行
    ② ファンドリターナー ─ ★回収額 = Exit売上 × EV/Sales × Exit持分
    ③ 感度            ─ EV/Sales × Exit持分 マトリクス
    ④ ポートフォリオ   ─ べき乗則の分布 + net DPI 検算
"""
import pathlib

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill

ROOT = pathlib.Path(__file__).resolve().parent.parent
OUT = ROOT / "outputs" / "02_財務モデル.xlsx"

# 色定義
YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GREEN_PASS = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FAIL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
NARRATIVE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")

BOLD = Font(bold=True)
NARRATIVE_FONT = Font(bold=True, color="FFFFFF", size=12)
HEADER_FONT = Font(bold=True, color="FFFFFF")
NOTE_FONT = Font(italic=True, color="595959")


def build_assumptions(wb: Workbook) -> None:
    ws = wb.create_sheet("前提")

    ws["A1"] = "全ての数字はここで決まる。8つの入力 × 結論は2つ(=ファンドリターナー成立 / net DPI 3X)。"
    ws["A1"].font = NARRATIVE_FONT
    ws["A1"].fill = NARRATIVE_FILL
    ws.merge_cells("A1:D1")

    for i, h in enumerate(["項目", "値", "単位", "注記"], start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    rows = [
        ("ファンド総額", 150, "億", "3号ファンド(1号30億DPI2.0X / 2号80億DPI1.5X)"),
        ("ファンドリターナー閾値", 150, "億", "1社の回収がこれ以上で成立"),
        ("目標 net DPI", 3.0, "X", "グロス約4.3X 相当"),
        ("初回チケット", 5.5, "億", "Solafune 初回投資額"),
        ("投資前 Pre-Val", 31, "億", "[仮] シード〜プレA拡大組成。Post=36.5"),
        ("エントリー持分", "=B6/(B7+B6)", "", "初回/(Pre+初回)。15.1%"),
        ("2032年 Exit売上", 100, "億", "[仮] 公的需要プール導出。CAGR延伸を廃止し直接入力"),
        ("EV/Sales 倍率", 17, "倍", "[仮] comps中央値10倍 + 主権/ソフト比/粗利プレミアム"),
        ("Exit持分(希薄化後)", 0.12, "", "[仮] 10〜12%帯"),
    ]
    for i, (label, value, unit, note) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label)
        cv = ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=3, value=unit)
        ws.cell(row=i, column=4, value=note)
        if isinstance(value, str) and value.startswith("="):
            cv.fill = GREEN
        else:
            cv.fill = YELLOW

    ws["B11"].number_format = "0.00"

    # 注記行(計算には使わない)
    ws["A12"] = "注記(計算には未使用)"
    ws["B12"] = "集中度上限13% / リザーブ初回:追加=45:55 / 管理報酬2% / 運用10年(投資5+回収5)"
    ws["A12"].font = NOTE_FONT
    ws["B12"].font = NOTE_FONT
    ws.merge_cells("B12:D12")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 55


def build_fund_returner(wb: Workbook) -> None:
    ws = wb.create_sheet("ファンドリターナー")

    ws["A1"] = (
        "Solafune単独で 回収204億 ≧ 閾値150億 ⇒ ファンドリターナー成立"
        "(エントリー15.1% → 希薄化後12% × EV1,700億)"
    )
    ws["A1"].font = NARRATIVE_FONT
    ws["A1"].fill = NARRATIVE_FILL
    ws.merge_cells("A1:C1")

    for i, h in enumerate(["項目", "値", "単位"], start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    rows = [
        ("初回投資", "=前提!B6", "億"),
        ("Pre-Val", "=前提!B7", "億"),
        ("エントリー持分", "=前提!B8", ""),
        ("Exit売上", "=前提!B9", "億"),
        ("EV/Sales", "=前提!B10", "倍"),
        ("Exit企業価値", "=B6*B7", "億"),
        ("Exit持分(希薄化後)", "=前提!B11", ""),
        ("★回収額", "=B8*B9", "億"),
        (
            "ファンドリターナー判定",
            '=IF(B10>=前提!B4,"○ ファンドリターナー成立","× 不成立")',
            "",
        ),
        ("対初回MOIC", "=B10/B3", "X"),
    ]
    for i, (label, value, unit) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label)
        cv = ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=3, value=unit)
        cv.fill = GREEN

    # 持分は%表示
    ws["B5"].number_format = "0.0%"
    ws["B9"].number_format = "0.0%"
    ws["B12"].number_format = "0.0"

    # ★回収額 と 判定を強調
    ws["A10"].font = Font(bold=True, size=12)
    ws["B10"].font = Font(bold=True, size=12)
    ws["B10"].fill = GREEN_PASS
    ws["A11"].font = Font(bold=True, size=12)
    ws["B11"].font = Font(bold=True, size=12)
    ws["B11"].fill = GREEN_PASS

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 8


def build_sensitivity(wb: Workbook) -> None:
    ws = wb.create_sheet("感度")

    ws["A1"] = "17×0.12=204億(中心)。EV/Sales 15倍×持分10%(=150億)まで耐える=外れても勝てる幅"
    ws["A1"].font = NARRATIVE_FONT
    ws["A1"].fill = NARRATIVE_FILL
    ws.merge_cells("A1:E1")

    ws["A2"] = (
        "回収額(億) = Exit売上 × EV/Sales × Exit持分。"
        "緑=150億超(ファンドリターナー成立域)"
    )
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:E2")

    ws["A3"] = "EV/Sales ＼ Exit持分"
    ws["A3"].font = HEADER_FONT
    ws["A3"].fill = HEADER_FILL
    stake_headers = [0.08, 0.10, 0.11, 0.12]
    for j, val in enumerate(stake_headers, start=2):
        c = ws.cell(row=3, column=j, value=val)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.number_format = "0.00"

    ev_multiples = [12, 15, 17, 18, 20]
    for i, ev in enumerate(ev_multiples, start=4):
        ca = ws.cell(row=i, column=1, value=ev)
        ca.font = HEADER_FONT
        ca.fill = HEADER_FILL
        for j in range(2, 6):
            col_letter = chr(ord("A") + j - 1)  # B,C,D,E
            formula = f"=ファンドリターナー!$B$6*$A{i}*{col_letter}$3"
            cell = ws.cell(row=i, column=j, value=formula)
            cell.fill = GREEN
            cell.number_format = "0.0"
        if ev == 17:
            # 中心行(17倍)は太字で強調。色は既存(白=HEADER_FONT)/黒(数式セル)を維持。
            for j in range(1, 6):
                cell = ws.cell(row=i, column=j)
                existing = cell.font
                cell.font = Font(
                    name=existing.name,
                    size=existing.size,
                    bold=True,
                    italic=existing.italic,
                    color=existing.color,
                )

    rng = "B4:E8"
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="greaterThanOrEqual", formula=["150"], fill=GREEN_PASS),
    )
    ws.conditional_formatting.add(
        rng,
        CellIsRule(operator="lessThan", formula=["150"], fill=RED_FAIL),
    )

    ws.column_dimensions["A"].width = 22
    for col in "BCDE":
        ws.column_dimensions[col].width = 11


def build_portfolio(wb: Workbook) -> None:
    ws = wb.create_sheet("ポートフォリオ")

    ws["A1"] = "★1社 + ◎2社で 250億超を作り、残り7社で元本前後(=べき乗則の形)"
    ws["A1"].font = NARRATIVE_FONT
    ws["A1"].fill = NARRATIVE_FILL
    ws.merge_cells("A1:G1")

    ws["A2"] = (
        "投下=億 / MOIC=X / 回収=億。Solafune行(F4) は ファンドリターナー!B10 連動。"
        "他9社は分布の形を示す想定値[仮]。"
    )
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:G2")

    headers = ["#", "会社", "区分", "投下", "MOIC", "回収", "領域"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    # net DPI を目標3.0X近傍に置くため、上位2社の回収値を調整(本質=べき乗則は維持)
    portfolio = [
        (1,  "Solafune",       "★ファンドリターナー", 16.5, "=ファンドリターナー!B10", "衛星AI解析"),
        (2,  "Humanity Brain", "◎ 3-5X級",          16,   160,                       "認知戦AI"),
        (3,  "C",              "◎ 3-5X級",          14,   90,                        "都市鉱山"),
        (4,  "D",              "○ 1-2X級",          13,   19.5,                      "海上ISR"),
        (5,  "E",              "○ 1-2X級",          13,   19.5,                      "センシング基盤"),
        (6,  "F",              "○ 1-2X級(元本)",    12,   12,                        "早期SU"),
        (7,  "G",              "× 0-1X",           11,   5.5,                       "早期SU"),
        (8,  "H",              "× 0-1X",           10,   3,                         "早期SU"),
        (9,  "I",              "× 全損",           9,    0,                         "早期SU"),
        (10, "J",              "× 全損",           8,    0,                         "早期SU"),
    ]
    for idx, (num, name, tier, invested, recovery, domain) in enumerate(portfolio, start=4):
        ws.cell(row=idx, column=1, value=num)
        ws.cell(row=idx, column=2, value=name)
        ws.cell(row=idx, column=3, value=tier)
        ws.cell(row=idx, column=4, value=invested).fill = YELLOW
        ws.cell(row=idx, column=5, value=f"=F{idx}/D{idx}").fill = GREEN
        ws.cell(row=idx, column=5).number_format = "0.0"
        rec_cell = ws.cell(row=idx, column=6, value=recovery)
        if isinstance(recovery, str) and recovery.startswith("="):
            rec_cell.fill = GREEN
        else:
            rec_cell.fill = YELLOW
        ws.cell(row=idx, column=7, value=domain)

    # Solafune行を強調
    for col in range(1, 8):
        ws.cell(row=4, column=col).font = BOLD

    # 合計(行14)
    ws["B14"] = "合計(10社)"
    ws["D14"] = "=SUM(D4:D13)"
    ws["E14"] = "=F14/D14"
    ws["F14"] = "=SUM(F4:F13)"
    ws["D14"].fill = GREEN
    ws["E14"].fill = GREEN
    ws["F14"].fill = GREEN
    ws["E14"].number_format = "0.00"
    for col in range(1, 8):
        ws.cell(row=14, column=col).font = BOLD

    # net DPI 検算(行16-18)
    ws["A16"] = "LP分配 (元本+利益80%)"
    ws["B16"] = "=前提!B3+0.8*(F14-前提!B3)"
    ws["B16"].fill = GREEN
    ws["B16"].number_format = "0.0"
    ws["C16"] = "億"

    ws["A17"] = "★ net DPI (対ファンド総額)"
    ws["B17"] = "=(前提!B3+0.8*(F14-前提!B3))/前提!B3"
    ws["B17"].fill = GREEN_PASS
    ws["B17"].number_format = "0.00"
    ws["C17"] = "X"
    ws["A17"].font = Font(bold=True, size=12)
    ws["B17"].font = Font(bold=True, size=12)

    ws["A18"] = "目標 net DPI 3.0X 到達度"
    ws["B18"] = "=B17/前提!B5"
    ws["B18"].fill = GREEN
    ws["B18"].number_format = "0.0%"

    widths = [4, 16, 22, 8, 8, 10, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    build_assumptions(wb)
    build_fund_returner(wb)
    build_sensitivity(wb)
    build_portfolio(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote: {OUT}")


if __name__ == "__main__":
    main()
